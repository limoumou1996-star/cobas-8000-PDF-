# processor.py (版本 2.0 - 支持多文件处理)

import re
import pdfplumber
import openpyxl
from io import BytesIO

# --- (这里是您所有的核心处理函数，保持不变) ---
# --- (函数 clean_row_data, process_pdf_text, create_raw_sheet, 等... 一直到 create_parsed_sheet) ---
def clean_row_data(columns):
    cleaned_columns = []
    for column in columns:
        if column not in ('*', '+'):
            cleaned_columns.append(column)
    return cleaned_columns

def process_pdf_text(text):
    lines = text.splitlines()
    processed_lines = []
    for line in lines:
        if (line.startswith("Result List") or
            "Data Type:" in line or
            "S. Type" in line or
            "Comment Operator ID" in line or
            "Test Result Alarm" in line or
            re.match(r"^\d{8}$", line) or
            line.strip() == ""):
            continue
        processed_lines.append(line)
    return processed_lines

def create_raw_sheet(workbook, lines, sheet_name):
    raw_sheet = workbook.create_sheet(sheet_name[:31])
    for row_idx, line in enumerate(lines, 1):
        columns = line.split()
        cleaned_columns = clean_row_data(columns)
        for col_idx, column_value in enumerate(cleaned_columns, 1):
            raw_sheet.cell(row=row_idx, column=col_idx, value=column_value)

def clean_line(line):
    return not (line.startswith("Result List") or
                "Data Type:" in line or
                "S. Type" in line or
                "Comment Operator ID" in line or
                "Test Result Alarm" in line or
                "System Name:" in line or
                "R2" in line or
                "R3" in line or
                "ProCell Lot ID" in line or
                re.match(r"^\d{8}$", line) or
                line.strip() == "")

def extract_numeric_field(value):
    try:
        float(value.replace(',', ''))
        return value
    except ValueError:
        return ""

def convert_to_number(value):
    try:
        return float(value.replace(',', ''))
    except ValueError:
        return value

def split_result_field(value):
    match = re.match(r"([-+]?[0-9]*\.?[0-9]+)(.*)", value)
    if match:
        return match.groups()
    return value, ""

def is_warning_field(value):
    return len(value) > 4 and '-' not in value and extract_numeric_field(value) == ""

def is_valid_reagent_position(part):
    if part.isdigit():
        number = int(part)
        return 1 <= number <= 60 and part == str(number)
    return False

def process_sample_data(sample, parsed_sheet, row_num, file_name):
    sample_info = {
        "样本序列号": "", "样本号": "", "样本类型": "", "架子号": "", "检测项目": "",
        "检测结果": "", "结果报警": "", "检测模块": "", "试剂位置": "", "稀释信息": "", "备注信息": "",
        "检测日期": "", "上机时间": "", "试剂批号": "", "试剂盒号": "", "文件名称": file_name
    }
    result_lines = []
    rack_id_pattern = r"^[A-Za-z]?\d{5}-\d{1,5}$"
    date_pattern = r"^\d{1,4}/\d{1,4}/\d{1,4}$"
    time_pattern = r"^\d{2}:\d{2}:\d{2}$"
    def capture_sample_info(line, sample_info):
        columns = line.split()
        sample_info["样本类型"] = columns[0]
        if len(columns) > 2 and re.match(rack_id_pattern, columns[1]):
            sample_info["架子号"] = columns[1]
        elif len(columns) > 3 and re.match(rack_id_pattern, columns[2]):
            sample_info["架子号"] = columns[2]
            sample_info["样本序列号"] = convert_to_number(columns[1])
        for col in columns:
            if re.match(date_pattern, col):
                sample_info["检测日期"] = col
            elif re.match(time_pattern, col):
                sample_info["上机时间"] = col
    def capture_additional_info(line, sample_info):
        columns = line.split()
        if len(columns) == 3: sample_info["备注信息"] = "-"
        elif len(columns) == 2: sample_info["备注信息"] = columns[0]
        elif ":" in line:
            parts = line.split(":")
            if len(parts) > 1: sample_info["备注信息"] = parts[1].strip().split()[0]
    def process_result_lines(start_index):
        result_line_index = start_index
        while result_line_index < len(sample):
            line = sample[result_line_index].strip()
            if re.match(r"(Ser/Pl|Urine|CSF)", line): break
            parts = line.split()
            cleaned_parts = clean_row_data(parts)
            for k, part in enumerate(cleaned_parts):
                if part in ('Samp.B', 'Samp.C', 'Samp.S', 'Reag.S', 'Reag.F', 'Reag.H', 'SLLD.N', 'SLLD.E', 'Over.E', 'Cal.E'):
                    cleaned_parts.insert(k, "-9999")
                    break
            if len(cleaned_parts) >= 2:
                if (len(cleaned_parts) > 2 and (cleaned_parts[1] in ['2', '3', '4', '5', 'K', 'NA', 'CL', 'E', 'R', '72-4', '21-1', '12-5', '15-3', 'II', 'III', '801', 'IV', '602', 'new']) and extract_numeric_field(cleaned_parts[2])):
                    cleaned_parts[0] = cleaned_parts[0] + ' ' + cleaned_parts[1]
                    cleaned_parts.pop(1)
                elif (len(cleaned_parts) > 3 and (cleaned_parts[1] in ['2', '3', '4', '5', 'K', 'NA', 'CL', 'E', 'R', '72-4', '21-1', '12-5', '15-3', 'II', 'III', '801', 'IV', '602', 'new']) and (cleaned_parts[2] in ['2', '3', '4', '5', 'K', 'NA', 'CL', 'E', 'R', '72-4', '21-1', '12-5', '15-3', 'II', 'III', '801', 'IV', '602', 'new']) and extract_numeric_field(cleaned_parts[3])):
                    cleaned_parts[0] = cleaned_parts[0] + ' ' + cleaned_parts[1] + ' ' + cleaned_parts[2]
                    cleaned_parts.pop(1); cleaned_parts.pop(1)
                result_info = {"检测项目": cleaned_parts[0], "检测结果": convert_to_number(cleaned_parts[1]), "结果报警": "", "检测模块": "", "试剂位置": "", "稀释信息": "", "试剂批号": "", "试剂盒号": ""}
                for i in range(2, len(cleaned_parts)):
                    part = cleaned_parts[i]
                    if part in ('Dec', 'Inc', 'Pre'): result_info["稀释信息"] = part; continue
                    split_numeric = re.match(r"^(\d+)(\D.*)$", part)
                    if split_numeric: result_info["稀释信息"] = split_numeric.group(1); result_info["检测模块"] = split_numeric.group(2); continue
                    if re.search(r"-[AB12]$|ISE|I1|AU", part): result_info["检测模块"] = part; break
                    if is_warning_field(part): result_info["结果报警"] = part
                if result_line_index < len(sample):
                    next_parts = sample[result_line_index].split()
                    cleaned_next_parts = clean_row_data(next_parts)
                    if cleaned_next_parts:
                        pos_candidates = [part for idx, part in enumerate(reversed(cleaned_next_parts)) if (is_valid_reagent_position(part) or re.match(r'[AB]-\d{1,2}', part)) and idx != 0]
                        if pos_candidates: result_info["试剂位置"] = pos_candidates[0]
                    if len(cleaned_next_parts) >= 5:
                        result_info["试剂盒号"] = extract_numeric_field(cleaned_next_parts[-1])
                        result_info["试剂批号"] = extract_numeric_field(cleaned_next_parts[-2])
                result_lines.append(result_info)
            result_line_index += 1
        return result_line_index
    i = 0
    while i < len(sample):
        line = sample[i].strip()
        if re.match(r"(Ser/Pl|Urine|CSF)", line):
            capture_sample_info(line, sample_info)
            grabbed_sample_id = re.search(r"ID\s*:\s*(\S+)|Lot\s*:\s*(\S+)", line)
            if grabbed_sample_id:
                sample_id = grabbed_sample_id.group(1) if grabbed_sample_id.group(1) else grabbed_sample_id.group(2)
                if sample_id and not re.match(date_pattern, sample_id): sample_info["样本号"] = sample_id
            if i + 1 < len(sample): capture_additional_info(sample[i + 1].strip(), sample_info)
            i += 2
            i = process_result_lines(i)
            continue
        i += 1
    for result in result_lines:
        parsed_sheet.append([
            sample_info["样本序列号"], sample_info["样本号"], sample_info["样本类型"], sample_info["架子号"], result["检测项目"], result["检测结果"], result["结果报警"],
            result["检测模块"], result["试剂位置"], result["稀释信息"], sample_info["备注信息"], sample_info["检测日期"], sample_info["上机时间"], result["试剂批号"], result["试剂盒号"], sample_info["文件名称"]
        ])
        row_num += 1
    return row_num

def create_parsed_sheet(workbook, raw_lines, sheet_name, file_name, from_folder=False):
    if from_folder: sheet_name = f"解析_{sheet_name[-28:]}"[:31]
    else: sheet_name = f"解析_{sheet_name[-28:]}"[:31]
    if sheet_name not in workbook.sheetnames:
        parsed_sheet = workbook.create_sheet(sheet_name, 0)
        headers = ["样本序列号", "样本号", "样本类型", "架子号", "检测项目", "检测结果", "结果报警", "检测模块", "试剂位置", "稀释信息", "备注信息", "检测日期", "上机时间", "试剂批号", "试剂盒号", "文件名称"]
        parsed_sheet.append(headers)
        parsed_sheet.freeze_panes = "A2"
    else: parsed_sheet = workbook[sheet_name]
    row_num = parsed_sheet.max_row + 1 if parsed_sheet.max_row > 1 else 2
    sample_data = []
    def process_existing_sample():
        nonlocal row_num, sample_data
        if sample_data:
            row_num = process_sample_data(sample_data, parsed_sheet, row_num, file_name)
            sample_data = []
    i = 0
    while i < len(raw_lines):
        line = raw_lines[i]
        if clean_line(line):
            if re.match(r"(Ser/Pl|Urine|CSF)", line): process_existing_sample()
            sample_data.append(line)
        i += 1
    process_existing_sample()
    return workbook

# --- (这是新的主函数，用于处理多个文件) ---
def process_multiple_pdfs_to_excel_bytes(files_data):
    """
    接收一个包含多个PDF文件信息（文件名和内容）的列表，
    处理后返回一个合并了所有结果的Excel文件的字节内容。
    """
    try:
        # 1. 创建一个总的Excel工作簿
        workbook = openpyxl.Workbook()
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # 2. 循环处理从网页传来的每一个文件
        # files_data 是一个列表，每个元素是 {'name': '文件名', 'content': 文件字节内容}
        for file_info in files_data:
            file_name = file_info['name']
            pdf_content = file_info['content']

            # 从内存中读取PDF内容
            text = ""
            with pdfplumber.open(BytesIO(pdf_content)) as pdf:
                for page in pdf.pages:
                    if page.extract_text():
                        text += page.extract_text()
            
            # 调用你已有的文本处理函数
            processed_lines = process_pdf_text(text)

            # 将解析结果添加到总的工作簿中
            # 注意：我们将所有解析结果都放在一个名为"解析_结果汇总"的sheet里
            create_parsed_sheet(workbook, processed_lines, "结果汇总", file_name, from_folder=True)

        # 3. 将最终的Excel文件保存到内存中
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.getvalue()

    except Exception as e:
        print(f"处理PDF时发生错误: {e}")
        return None